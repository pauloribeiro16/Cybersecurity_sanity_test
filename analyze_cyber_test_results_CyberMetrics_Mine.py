# analyze_test_summary.py

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import CellIsRule

# --- CONFIGURAÇÕES ---
INPUT_JSON_FILE = "Results_Teste_3.json"
OUTPUT_EXCEL_FILE = "Test_Summary_Report.xlsx"

# Mapeamento de cores para a formatação condicional
COLOR_MAPPING = {
    "CORRECT": "C6EFCE",    # Verde claro
    "INCORRECT": "FFC7CE",  # Vermelho claro
    "INCOMPLETE": "FFEB9C"  # Amarelo claro
}

# As categorias das perguntas na ordem em que foram executadas.
# Esta ordem deve corresponder a "Question_1", "Question_2", etc. no JSON.
CATEGORIES_IN_ORDER = [
    "Basic Definition",
    "Frameworks",
    "Security Controls",
    "Vulnerabilities",
    "Threat Modeling",
    "Regulations/Directives",
    "Basic Principles"
]

def build_summary_from_details(data: dict) -> pd.DataFrame:
    """
    Processa os detalhes do JSON para criar um DataFrame de resumo pivotado.
    """
    flat_data = []
    # Itera sobre os modelos e as suas respetivas perguntas
    for model_name, questions in data.get('details', {}).items():
        # Ordena as perguntas para garantir que correspondem à ordem das categorias
        sorted_q_keys = sorted(questions.keys(), key=lambda x: int(x.split('_')[1]))
        
        for i, q_key in enumerate(sorted_q_keys):
            if i < len(CATEGORIES_IN_ORDER):
                q_details = questions[q_key]
                flat_data.append({
                    "Model": model_name,
                    "Category": CATEGORIES_IN_ORDER[i],
                    "Evaluation": q_details.get("evaluation", "N/A").upper()
                })

    if not flat_data:
        return pd.DataFrame()

    # Cria o DataFrame e pivota para ter Modelos como linhas e Categorias como colunas
    df = pd.DataFrame(flat_data)
    pivot_df = df.pivot_table(index='Model', columns='Category', values='Evaluation', aggfunc='first')
    
    # Garante que a ordem das colunas de categoria está correta
    pivot_df = pivot_df[CATEGORIES_IN_ORDER]

    # --- Calcula as colunas de estatísticas ---
    total_questions = len(CATEGORIES_IN_ORDER)
    pivot_df["Correct"] = pivot_df.apply(lambda row: (row == 'CORRECT').sum(), axis=1)
    pivot_df["Incomplete"] = pivot_df.apply(lambda row: (row == 'INCOMPLETE').sum(), axis=1)
    pivot_df["Incorrect"] = pivot_df.apply(lambda row: (row == 'INCORRECT').sum(), axis=1)
    
    # Calcula a pontuação conforme solicitado (Corretas * 2)
    pivot_df["Score"] = pivot_df["Correct"] * 2
    
    # Calcula a percentagem de respostas corretas
    pivot_df["Correct %"] = pivot_df["Correct"] / total_questions
    
    # Reordena as colunas para corresponder à imagem
    final_order = CATEGORIES_IN_ORDER + ["Score", "Correct %", "Correct", "Incomplete", "Incorrect"]
    
    return pivot_df[final_order]


def apply_summary_formatting(worksheet, num_categories: int):
    """Aplica toda a formatação necessária à folha de resumo."""
    header_font = Font(bold=True, name='Calibri', size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid") # Azul escuro do Office
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Formata o cabeçalho
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Formata as células de dados
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = center_alignment
            if cell.column > num_categories + 1: # Colunas de estatísticas
                cell.font = Font(bold=True)
            if cell.column == num_categories + 3: # Coluna 'Correct %'
                cell.number_format = '0.0%'

    # Ajusta as larguras das colunas
    worksheet.column_dimensions['A'].width = 30  # Model
    for i in range(num_categories):
        col_letter = chr(ord('B') + i)
        worksheet.column_dimensions[col_letter].width = 22
    for i in range(5):
        col_letter = chr(ord('B') + num_categories + i)
        worksheet.column_dimensions[col_letter].width = 12

    # Aplica formatação condicional de cores para as categorias
    range_to_format = f"B2:{chr(ord('A') + num_categories)}{worksheet.max_row}"
    for eval_text, color_hex in COLOR_MAPPING.items():
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        font = Font(color="006100" if eval_text == "CORRECT" else ("9C5700" if eval_text == "INCOMPLETE" else "9C0006"))
        rule = CellIsRule(operator="equal", formula=[f'"{eval_text}"'], fill=fill, font=font)
        worksheet.conditional_formatting.add(range_to_format, rule)


def main():
    """Função principal para gerar o relatório Excel."""
    print(f"Lendo ficheiro de resultados: {INPUT_JSON_FILE}")
    try:
        with open(INPUT_JSON_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"ERRO: Ficheiro '{INPUT_JSON_FILE}' não encontrado.")
        return
    except json.JSONDecodeError:
        print(f"ERRO: Ficheiro '{INPUT_JSON_FILE}' não é um JSON válido.")
        return

    print("A processar os resultados e a construir a tabela de resumo...")
    summary_df = build_summary_from_details(data)

    if summary_df.empty:
        print("Não foram encontrados dados de 'details' para processar.")
        return

    # Cria o livro e a folha de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Model Test Summary"

    # Escreve o DataFrame para a folha
    for r in dataframe_to_rows(summary_df, index=True, header=True):
        ws.append(r)

    print("A aplicar formatação ao ficheiro Excel...")
    apply_summary_formatting(ws, len(CATEGORIES_IN_ORDER))

    # Salva o ficheiro Excel
    try:
        wb.save(OUTPUT_EXCEL_FILE)
        print(f"\nSucesso! O relatório de resumo foi guardado em '{OUTPUT_EXCEL_FILE}'.")
    except PermissionError:
        print(f"\nERRO: Permissão negada. O ficheiro '{OUTPUT_EXCEL_FILE}' pode estar aberto. Por favor, feche-o e tente novamente.")
    except Exception as e:
        print(f"\nERRO: Ocorreu um erro ao guardar o ficheiro Excel: {e}")

if __name__ == "__main__":
    main()