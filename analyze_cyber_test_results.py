# analyze_cyber_test_results.py

import json
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo

# --- CONFIGURAÇÕES ---
INPUT_JSON_FILE = "Results.json"
OUTPUT_EXCEL_FILE = "Cybersecurity_Test_Analysis_v2.xlsx"


SCORE_MAPPING = {"Correct": 2, "Incomplete": 1, "Incorrect": 0, "N.A.": 0}
COLOR_MAPPING = {"Correct": "C6EFCE", "Incomplete": "FFEB9C", "Incorrect": "FFC7CE"}
CHART_SCORE_COLOR = "4472C4" 

# --- FUNÇÕES AUXILIARES ---

def get_model_family(model_name: str) -> str:
    # (Sem alterações)
    if ':' in model_name: return model_name.split(':')[0]
    if '-' in model_name: return model_name.split('-')[0]
    return model_name

# <<< ALTERAÇÃO: Processar os novos campos do JSON
def process_results(data: dict) -> tuple:
    summary_data_models = []
    details_by_family = defaultdict(list)
    sorted_model_names = sorted(data.keys())
    
    if not sorted_model_names:
        return [], {}, []

    first_model_name = next(iter(sorted_model_names))
    categories = list(data[first_model_name].keys())
    
    for model_name in sorted_model_names:
        results = data[model_name]
        summary_data_models.append(model_name)
        model_family = get_model_family(model_name)
        for category in categories:
            result = results.get(category, {})
            details_by_family[model_family].append({
                "Model": model_name,
                "Category": category,
                "Evaluation": result.get("evaluation", "Missing"),
                "Prompt": result.get("prompt", ""),
                "Expected Answer": result.get("expected_short_answer", ""),
                "LLM Response": result.get("llm_response", ""),
                "Analysis": result.get("analysis", "")
            })
            
    return summary_data_models, details_by_family, categories

# <<< ALTERAÇÃO: Formatar as novas colunas
def apply_formatting(worksheet, is_summary=False):
    """Aplica formatação genérica (alinhamento, fonte) a uma folha."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Formatar cabeçalho
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Formatar dados
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            # Colunas com texto longo são alinhadas à esquerda/topo
            if cell.column_letter in ['D', 'E', 'F', 'G']: # Prompt, Expected, LLM Response, Analysis
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment

    # Definir larguras
    if is_summary:
        worksheet.column_dimensions['A'].width = 35
        for col in worksheet.iter_cols(min_col=2, max_col=worksheet.max_column):
             worksheet.column_dimensions[col[0].column_letter].width = 20
    else: # Folha de detalhe
        worksheet.column_dimensions['A'].width = 35  # Model
        worksheet.column_dimensions['B'].width = 25  # Category
        worksheet.column_dimensions['C'].width = 15  # Evaluation
        worksheet.column_dimensions['D'].width = 60  # Prompt
        worksheet.column_dimensions['E'].width = 80  # Expected Answer
        worksheet.column_dimensions['F'].width = 80  # LLM Response
        worksheet.column_dimensions['G'].width = 80  # Analysis


def apply_conditional_formatting(worksheet, range_str):
    """Aplica formatação condicional de cores a um intervalo de células."""
    for evaluation, color_hex in COLOR_MAPPING.items():
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        rule = CellIsRule(operator="equal", formula=[f'"{evaluation}"'], fill=fill)
        worksheet.conditional_formatting.add(range_str, rule)

def create_charts(workbook, summary_sheet_name, models, categories):
    # (Sem alterações aqui, esta função já é robusta o suficiente)
    if "Dashboard" in workbook.sheetnames:
        dashboard_ws = workbook["Dashboard"]
    else:
        dashboard_ws = workbook.create_sheet("Dashboard", 0)

    summary_ws = workbook[summary_sheet_name]
    num_models = len(models)
    score_col_letter = chr(ord('B') + len(categories))
    
    # Gráfico 1: Pontuação Total
    chart1 = BarChart()
    chart1.type = "bar"
    chart1.title = "Total Score by Model"
    chart1.style = 11
    chart1.y_axis.title = 'Model'
    chart1.x_axis.title = 'Total Score'
    data = Reference(summary_ws, min_col=ord(score_col_letter) - ord('A') + 1, min_row=2, max_row=num_models + 1)
    cats = Reference(summary_ws, min_col=1, min_row=2, max_row=num_models + 1)
    chart1.add_data(data, titles_from_data=False)
    chart1.set_categories(cats)
    chart1.legend = None
    series1 = chart1.series[0]
    series1.graphicalProperties.solidFill = CHART_SCORE_COLOR
    chart1.y_axis.scaling.orientation = "maxMin"
    dashboard_ws.add_chart(chart1, "B2")
    chart1.width = 15
    chart1.height = 10
    
    # Gráfico 2: Distribuição
    chart2 = BarChart()
    chart2.type = "bar"
    chart2.grouping = "stacked"
    chart2.overlap = 100
    chart2.title = "Evaluation Distribution by Model"
    chart2.style = 11
    chart2.y_axis.title = 'Model'
    chart2.x_axis.title = 'Number of Test Cases'
    chart2.set_categories(cats)
    eval_types_for_chart = ["Correct", "Incomplete", "Incorrect"]
    for i, eval_type in enumerate(eval_types_for_chart):
        count_col_offset = len(categories) + 3 + i
        count_col = ord('A') + count_col_offset
        values = Reference(summary_ws, min_col=count_col - ord('A') + 1, min_row=2, max_row=num_models + 1)
        series = Series(values, title=eval_type)
        series.graphicalProperties.solidFill = COLOR_MAPPING.get(eval_type, "FFFFFF")
        chart2.append(series)
    chart2.y_axis.scaling.orientation = "maxMin"
    dashboard_ws.add_chart(chart2, "K2")
    chart2.width = 15
    chart2.height = 10


# --- FUNÇÃO PRINCIPAL ---

def create_analysis_excel():
    print(f"Lendo ficheiro de resultados: {INPUT_JSON_FILE}")
    try:
        with open(INPUT_JSON_FILE, 'r', encoding='utf-8') as f: data = json.load(f)
    except FileNotFoundError: print(f"ERRO: Ficheiro '{INPUT_JSON_FILE}' não encontrado."); return
    except json.JSONDecodeError: print(f"ERRO: Ficheiro '{INPUT_JSON_FILE}' não é um JSON válido."); return
    
    print("Processando resultados...")
    summary_models, details_by_family, categories = process_results(data)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # --- CRIAR FOLHAS DE DETALHE (A "FONTE DA VERDADE") ---
    print("Criando folhas de detalhe por família...")
    for family_name, details_list in sorted(details_by_family.items()):
        sheet_name = family_name[:31]
        ws = wb.create_sheet(sheet_name)
        
        # <<< ALTERAÇÃO: Nova ordem de colunas para corresponder ao pedido
        df_family = pd.DataFrame(details_list)[[
            "Model", "Category", "Evaluation", "Prompt", 
            "Expected Answer", "LLM Response", "Analysis"
        ]]
        
        for r in dataframe_to_rows(df_family, index=False, header=True):
            ws.append(r)
        
        apply_formatting(ws)
        # A coluna de avaliação é agora a 'C'
        apply_conditional_formatting(ws, f"C2:C{ws.max_row}")
        print(f"  - Folha '{sheet_name}' criada. Edite as avaliações aqui (coluna C).")

    # --- CRIAR FOLHA DE SUMÁRIO COM RASTREABILIDADE ---
    print("Criando folha 'Summary' com fórmulas rastreáveis...")
    summary_ws = wb.create_sheet("Summary")
    
    summary_headers = ["Model"] + categories + ["Score", "Correct %", "Correct", "Incomplete", "Incorrect"]
    summary_ws.append(summary_headers)

    for row_idx, model_name in enumerate(summary_models, start=2):
        model_family = get_model_family(model_name)
        detail_sheet_name = f"'{model_family[:31]}'"
        
        row_data = [model_name]
        for col_idx, category in enumerate(categories, start=2):
            category_header_cell = f"{chr(ord('A') + col_idx - 1)}1"
            # <<< ALTERAÇÃO NA FÓRMULA DE RASTREABILIDADE
            # A fórmula agora tem de encontrar a Categoria (coluna B) e o Modelo (coluna A) e devolver a Avaliação (coluna C)
            # A fórmula INDEX/MATCH continua a ser perfeita para isto.
            # =INDEX('{sheet}'!$C:$C, MATCH(1, ('{sheet}'!$A:$A=$A{row})*('{sheet}'!$B:$B=B$1), 0))
            # Como a coluna C (Evaluation) não mudou de posição, a fórmula permanece a mesma e continua correta.
            formula = f'=INDEX({detail_sheet_name}!$C:$C, MATCH(1,({detail_sheet_name}!$A:$A=$A{row_idx})*({detail_sheet_name}!$B:$B={category_header_cell}),0))'
            row_data.append(formula)
        
        summary_ws.append(row_data)

    # Adicionar fórmulas de estatísticas (sem alterações, a lógica é a mesma)
    num_categories = len(categories)
    cat_range_letter_end = chr(ord('B') + num_categories - 1)
    score_col_letter = chr(ord('A') + num_categories + 1)
    correct_pct_col_letter = chr(ord('A') + num_categories + 2)
    correct_count_col_letter = chr(ord('A') + num_categories + 3)
    incomplete_count_col_letter = chr(ord('A') + num_categories + 4)
    incorrect_count_col_letter = chr(ord('A') + num_categories + 5)
    for row_idx in range(2, len(summary_models) + 2):
        cat_range = f"B{row_idx}:{cat_range_letter_end}{row_idx}"
        summary_ws[f"{score_col_letter}{row_idx}"] = f'=COUNTIF({cat_range},"Correct")*{SCORE_MAPPING["Correct"]} + COUNTIF({cat_range},"Incomplete")*{SCORE_MAPPING["Incomplete"]}'
        summary_ws[f"{correct_count_col_letter}{row_idx}"] = f'=COUNTIF({cat_range}, "Correct")'
        summary_ws[f"{incomplete_count_col_letter}{row_idx}"] = f'=COUNTIF({cat_range}, "Incomplete")'
        summary_ws[f"{incorrect_count_col_letter}{row_idx}"] = f'=COUNTIF({cat_range}, "Incorrect")'
        total_tests_formula = f'{correct_count_col_letter}{row_idx}+{incomplete_count_col_letter}{row_idx}+{incorrect_count_col_letter}{row_idx}'
        summary_ws[f"{correct_pct_col_letter}{row_idx}"] = f'=IF(({total_tests_formula})>0, {correct_count_col_letter}{row_idx}/({total_tests_formula}), 0)'
        summary_ws[f"{correct_pct_col_letter}{row_idx}"].number_format = '0.0%'

    apply_formatting(summary_ws, is_summary=True)
    apply_conditional_formatting(summary_ws, f"B2:{cat_range_letter_end}{len(summary_models) + 1}")
    
    print("  - Folha 'Summary' criada.")
    
    # --- CRIAR FOLHA DE GRÁFICOS ---
    print("Criando folha 'Dashboard' com gráficos...")
    create_charts(wb, "Summary", summary_models, categories)
    print("  - Gráficos criados.")

    # --- SALVAR O FICHEIRO ---
    try:
        wb.save(OUTPUT_EXCEL_FILE)
        print(f"\nSucesso! Análise guardada em '{OUTPUT_EXCEL_FILE}'.")
    except Exception as e:
        print(f"\nERRO: Ocorreu um erro ao gerar o ficheiro Excel: {e}")

if __name__ == "__main__":
    create_analysis_excel()