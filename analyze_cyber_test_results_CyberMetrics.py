# analyze_cyber_test_results.py

import json
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference, Series

# --- CONFIGURAÇÕES ---
# Altere este nome para o ficheiro de resultados que pretende analisar
INPUT_JSON_FILE = "Results.json" 
OUTPUT_EXCEL_FILE = f"Analysis_Report_{os.path.splitext(INPUT_JSON_FILE)[0]}.xlsx"

# Mapeamento de cores para formatação condicional
COLOR_MAPPING = {"CORRECT": "C6EFCE", "INCORRECT": "FFC7CE"}
CHART_ACCURACY_COLOR = "4472C4"  # Azul
CHART_CORRECT_COLOR = "86BC25"   # Verde
CHART_INCORRECT_COLOR = "C00000" # Vermelho


def process_detailed_results(data: dict) -> pd.DataFrame:
    """
    Transforma a secção 'details' do JSON numa lista simples para fácil conversão para DataFrame.
    """
    all_results = []
    
    # Itera sobre cada modelo na secção de detalhes
    for model_name, questions in data.items():
        # Itera sobre cada pergunta para esse modelo
        for question_id, details in questions.items():
            all_results.append({
                "Model": model_name,
                "Question ID": question_id,
                "Evaluation": details.get("evaluation", "N/A"),
                "LLM's Answer": details.get("extracted_answer", ""),
                "Correct Answer": details.get("expected_answer", ""),
                "Is Correct": details.get("is_correct", False),
                "Response Time (s)": float(details.get("response_time", 0.0)),
                "Full Question": details.get("full_prompt", ""),
                "Full LLM Response": details.get("full_llm_response", "")
            })
            
    return pd.DataFrame(all_results)

def apply_formatting(worksheet, column_widths: dict):
    """Aplica formatação genérica (cabeçalho, alinhamento, largura) a uma folha."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Formatar cabeçalho
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Formatar colunas
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            # Colunas com texto longo ficam alinhadas à esquerda/topo
            if cell.column_letter in ['H', 'I']: # Full Question, Full LLM Response
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment
    
    # Ajustar larguras das colunas
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width


def create_dashboard_charts(workbook, summary_df: pd.DataFrame):
    """Cria gráficos de barras na folha 'Dashboard' a partir do DataFrame de resumo."""
    if "Dashboard" not in workbook.sheetnames:
        dashboard_ws = workbook.create_sheet("Dashboard", 0)
    else:
        dashboard_ws = workbook["Dashboard"]

    summary_ws_name = "Summary"
    summary_ws = workbook[summary_ws_name]
    num_models = len(summary_df)

    # --- Gráfico 1: Precisão (Accuracy) por Modelo ---
    chart1 = BarChart()
    chart1.type = "col" # Colunas verticais
    chart1.title = "Model Accuracy"
    chart1.style = 10
    chart1.y_axis.title = 'Accuracy'
    chart1.y_axis.majorGridlines = None
    chart1.x_axis.title = 'Model'

    # Dados (coluna 'accuracy') e categorias (nomes dos modelos no índice)
    data = Reference(summary_ws, min_col=summary_df.columns.get_loc('accuracy') + 2, min_row=2, max_row=num_models + 1)
    cats = Reference(summary_ws, min_col=1, min_row=2, max_row=num_models + 1)
    
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.legend = None # Sem legenda para um único conjunto de dados

    # Cor da barra
    series1 = chart1.series[0]
    series1.graphicalProperties.solidFill = CHART_ACCURACY_COLOR
    
    dashboard_ws.add_chart(chart1, "B2")
    chart1.width = 15
    chart1.height = 12

    # --- Gráfico 2: Contagem de Respostas Corretas vs. Incorretas ---
    chart2 = BarChart()
    chart2.type = "bar"
    chart2.grouping = "stacked" # Barras empilhadas
    chart2.overlap = 100
    chart2.title = "Correct vs. Incorrect Answers"
    chart2.style = 10
    chart2.y_axis.title = 'Model'
    chart2.x_axis.title = 'Number of Answers'

    chart2.set_categories(cats) # Mesmas categorias (modelos)
    
    # Adicionar séries para 'correct' e 'incorrect'
    for col_name, color in [('correct', CHART_CORRECT_COLOR), ('incorrect', CHART_INCORRECT_COLOR)]:
        col_num = summary_df.columns.get_loc(col_name) + 2
        values = Reference(summary_ws, min_col=col_num, min_row=2, max_row=num_models + 1)
        series = Series(values, title=col_name.capitalize())
        series.graphicalProperties.solidFill = color
        chart2.append(series)
    
    chart2.y_axis.scaling.orientation = "maxMin" # Inverte a ordem do eixo Y
    dashboard_ws.add_chart(chart2, "L2")
    chart2.width = 15
    chart2.height = 12


# --- FUNÇÃO PRINCIPAL ---
def create_analysis_excel():
    print(f"Reading results from: {INPUT_JSON_FILE}")
    try:
        with open(INPUT_JSON_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"ERROR: File '{INPUT_JSON_FILE}' not found.")
        return
    except json.JSONDecodeError:
        print(f"ERROR: File '{INPUT_JSON_FILE}' is not a valid JSON.")
        return

    # --- 1. Processar a secção de 'summary' ---
    print("Processing summary data...")
    if 'summary' not in data:
        print("ERROR: 'summary' key not found in JSON file.")
        return
    summary_df = pd.DataFrame.from_dict(data['summary'], orient='index')
    # Converter a percentagem de string para float para ordenação
    summary_df['accuracy'] = summary_df['accuracy'].str.rstrip('%').astype(float) / 100.0
    summary_df = summary_df.sort_values(by='accuracy', ascending=False)
    
    # --- 2. Processar a secção de 'details' ---
    print("Processing detailed results...")
    if 'details' not in data:
        print("ERROR: 'details' key not found in JSON file.")
        return
    details_df = process_detailed_results(data['details'])

    # --- 3. Criar o ficheiro Excel ---
    wb = Workbook()
    wb.remove(wb.active) # Remove a folha padrão

    # --- Folha de Resumo ---
    print("Creating 'Summary' sheet...")
    summary_ws = wb.create_sheet("Summary", 1)
    for r in dataframe_to_rows(summary_df, index=True, header=True):
        summary_ws.append(r)
    # Formatação da folha de resumo
    summary_ws.column_dimensions['A'].width = 30 # Coluna do Modelo
    summary_ws['E2'].number_format = '0.00%' # Formato de percentagem
    apply_formatting(summary_ws, {'A': 35, 'B': 15, 'C': 15, 'D': 15, 'E': 15})
    
    # --- Folha de Detalhes ---
    print("Creating 'All Details' sheet...")
    details_ws = wb.create_sheet("All Details", 2)
    for r in dataframe_to_rows(details_df, index=False, header=True):
        details_ws.append(r)
    # Formatação condicional para a coluna 'Evaluation' (agora é a coluna C)
    fill_correct = PatternFill(start_color=COLOR_MAPPING["CORRECT"], end_color=COLOR_MAPPING["CORRECT"], fill_type="solid")
    fill_incorrect = PatternFill(start_color=COLOR_MAPPING["INCORRECT"], end_color=COLOR_MAPPING["INCORRECT"], fill_type="solid")
    details_ws.conditional_formatting.add(f"C2:C{details_ws.max_row}", CellIsRule(operator="equal", formula=['"CORRECT"'], fill=fill_correct))
    details_ws.conditional_formatting.add(f"C2:C{details_ws.max_row}", CellIsRule(operator="equal", formula=['"INCORRECT"'], fill=fill_incorrect))
    apply_formatting(details_ws, {
        'A': 35, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 20, 'H': 80, 'I': 80
    })

    # --- Folha de Dashboard com Gráficos ---
    print("Creating 'Dashboard' sheet with charts...")
    create_dashboard_charts(wb, summary_df)
    
    # --- Guardar Ficheiro ---
    try:
        wb.save(OUTPUT_EXCEL_FILE)
        print(f"\nSuccess! Analysis report saved to '{OUTPUT_EXCEL_FILE}'.")
    except Exception as e:
        print(f"\nERROR: Could not save the Excel file. It might be open. Error: {e}")

if __name__ == "__main__":
    create_analysis_excel()